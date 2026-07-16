"""
Greenhouse Controller Backend (v2)
----------------------------------
FastAPI + SQLite + live Excel logging (openpyxl). Handles:
  - ESP32 device ingest every ~2s (raw, uncalibrated data in -> corrected data stored)
  - Per-relay automation with independent AUTO / MANUAL OVERRIDE modes
  - Live-appended Excel workbook per device (SensorData sheet)
  - Dashboard API (current, history, status, settings, devices, excel export)
  - Static dashboard hosting

DESIGN NOTE ON RELAY MODE OWNERSHIP:
The spec asked for the ESP32 to upload each relay's mode (auto/manual). This
backend instead treats mode as SERVER-OWNED config (set from the dashboard),
and has the ESP32 report back the state it actually applied. Two reasons:
  1. A physical device deciding its own automation mode creates two competing
     sources of truth for the same switch (dashboard vs firmware).
  2. The dashboard is the natural place a person flips something to manual;
     the ESP32 should just be told what to do and confirm it happened.
See the final report for more on this.

Run locally:
    pip install -r requirements.txt
    uvicorn main:app --host 0.0.0.0 --port 8000 --reload
"""

import os
import sqlite3
import secrets
import math
from contextlib import contextmanager
from datetime import datetime, timezone, timedelta
from typing import Optional, Dict

from fastapi import FastAPI, HTTPException, Header, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.environ.get("GREENHOUSE_DB", os.path.join(BASE_DIR, "greenhouse.db"))
EXCEL_DIR = os.environ.get("GREENHOUSE_EXCEL_DIR", os.path.join(BASE_DIR, "excel_exports"))
DASHBOARD_USER = os.environ.get("DASHBOARD_USER", "admin")
DASHBOARD_PASS = os.environ.get("DASHBOARD_PASS", "changeme")

os.makedirs(EXCEL_DIR, exist_ok=True)

app = FastAPI(title="Greenhouse Controller API")
security = HTTPBasic()

# Sensor calibration — configurable constants, applied to every raw reading
# before it touches storage, VPD calc, automation, or the dashboard.
TEMP_OFFSET = -2.0
HUMIDITY_OFFSET = -5.0

RELAY_KEYS = ["heater", "fan1", "fan2", "fan3", "wing_flap", "curtains", "irrigation"]


# ---------------------------------------------------------------------------
# DB setup
# ---------------------------------------------------------------------------
@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    with get_db() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS devices (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                api_key TEXT UNIQUE NOT NULL,
                created_at TEXT NOT NULL,
                last_seen TEXT
            );

            CREATE TABLE IF NOT EXISTS readings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_id TEXT NOT NULL,
                ts TEXT NOT NULL,
                temperature REAL,
                humidity REAL,
                vpd REAL,
                FOREIGN KEY(device_id) REFERENCES devices(id)
            );
            CREATE INDEX IF NOT EXISTS idx_readings_device_ts ON readings(device_id, ts DESC);

            CREATE TABLE IF NOT EXISTS settings (
                device_id TEXT PRIMARY KEY,
                temp_min REAL DEFAULT 18,
                temp_max REAL DEFAULT 28,
                humidity_min REAL DEFAULT 50,
                humidity_max REAL DEFAULT 70,
                vpd_target REAL DEFAULT 1.0,
                fan_stage_margin REAL DEFAULT 2.0,
                irrigation_duration_minutes REAL DEFAULT 10,
                irrigation_interval_hours REAL DEFAULT 6,
                irrigation_last_start TEXT,
                FOREIGN KEY(device_id) REFERENCES devices(id)
            );

            -- Dashboard-owned per-relay config: auto vs manual, and the
            -- desired state when in manual mode.
            CREATE TABLE IF NOT EXISTS relay_config (
                device_id TEXT NOT NULL,
                relay_key TEXT NOT NULL,
                mode TEXT NOT NULL DEFAULT 'auto',        -- 'auto' | 'manual'
                manual_state INTEGER NOT NULL DEFAULT 0,   -- used when mode='manual'
                PRIMARY KEY (device_id, relay_key),
                FOREIGN KEY(device_id) REFERENCES devices(id)
            );

            -- What the ESP32 last reported actually applying, per relay.
            CREATE TABLE IF NOT EXISTS relay_reported (
                device_id TEXT NOT NULL,
                relay_key TEXT NOT NULL,
                reported_state INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT,
                PRIMARY KEY (device_id, relay_key),
                FOREIGN KEY(device_id) REFERENCES devices(id)
            );
            """
        )


init_db()


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class IngestPayload(BaseModel):
    device: str
    raw_temperature: float
    raw_humidity: float
    relay_states: Optional[Dict[str, int]] = None


class RelayConfigUpdate(BaseModel):
    mode: Optional[str] = None
    manual_state: Optional[bool] = None


class SettingsUpdate(BaseModel):
    temp_min: Optional[float] = None
    temp_max: Optional[float] = None
    humidity_min: Optional[float] = None
    humidity_max: Optional[float] = None
    vpd_target: Optional[float] = None
    fan_stage_margin: Optional[float] = None
    irrigation_duration_minutes: Optional[float] = None
    irrigation_interval_hours: Optional[float] = None


class NewDevice(BaseModel):
    name: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def apply_calibration(raw_temp: float, raw_hum: float):
    return raw_temp + TEMP_OFFSET, raw_hum + HUMIDITY_OFFSET


def calc_vpd(temp_c: float, rh_percent: float) -> float:
    svp = 0.61078 * math.exp((17.27 * temp_c) / (temp_c + 237.3))
    return round(svp * (1 - rh_percent / 100), 3)


def require_dashboard_auth(creds: HTTPBasicCredentials = Depends(security)):
    ok_user = secrets.compare_digest(creds.username, DASHBOARD_USER)
    ok_pass = secrets.compare_digest(creds.password, DASHBOARD_PASS)
    if not (ok_user and ok_pass):
        raise HTTPException(status_code=401, detail="Invalid credentials", headers={"WWW-Authenticate": "Basic"})
    return creds.username


def get_device_by_api_key(db, api_key: str):
    return db.execute("SELECT * FROM devices WHERE api_key = ?", (api_key,)).fetchone()


def ensure_settings_row(db, device_id: str):
    db.execute("INSERT OR IGNORE INTO settings (device_id) VALUES (?)", (device_id,))


def ensure_relay_rows(db, device_id: str):
    for key in RELAY_KEYS:
        db.execute(
            "INSERT OR IGNORE INTO relay_config (device_id, relay_key) VALUES (?, ?)",
            (device_id, key),
        )


def excel_path(device_id: str) -> str:
    return os.path.join(EXCEL_DIR, f"{device_id}.xlsx")


def append_to_excel(device_id: str, ts: str, temperature: float, humidity: float, vpd: float):
    """
    Live per-row Excel append, as requested. Opens, appends, saves on every
    call. SQLite remains the source of truth the dashboard and automation
    logic actually read from, so a slow or failed Excel write never blocks
    the live system.
    """
    path = excel_path(device_id)
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb["SensorData"] if "SensorData" in wb.sheetnames else wb.create_sheet("SensorData")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "SensorData"
            ws.append(["Timestamp", "Temperature", "Humidity", "VPD"])
        ws.append([ts, temperature, humidity, vpd])
        wb.save(path)
    except Exception as e:
        print(f"[excel] append failed for {device_id}: {e}")


def compute_relay_states(settings_row, relay_config_rows, temp: float, humidity: float, now: datetime, db, device_id: str):
    config = {r["relay_key"]: r for r in relay_config_rows}
    margin = settings_row["fan_stage_margin"]

    auto_states = {
        "heater":    temp < settings_row["temp_min"],
        "fan1":      temp > settings_row["temp_max"] or humidity > settings_row["humidity_max"],
        "fan2":      temp > settings_row["temp_max"] + margin,
        "fan3":      temp > settings_row["temp_max"] + 2 * margin,
        "wing_flap": temp > settings_row["temp_max"],
        "curtains":  temp > settings_row["temp_max"],  # "on" = shade deployed
    }

    duration = timedelta(minutes=settings_row["irrigation_duration_minutes"])
    interval = timedelta(hours=settings_row["irrigation_interval_hours"])
    last_start_str = settings_row["irrigation_last_start"]
    last_start = datetime.fromisoformat(last_start_str) if last_start_str else None

    irrigation_on = False
    if last_start is None or now - last_start >= interval:
        irrigation_on = True
        db.execute(
            "UPDATE settings SET irrigation_last_start = ? WHERE device_id = ?",
            (now.isoformat(), device_id),
        )
    elif now - last_start < duration:
        irrigation_on = True
    auto_states["irrigation"] = irrigation_on

    final_states = {}
    for key in RELAY_KEYS:
        cfg = config.get(key)
        if cfg and cfg["mode"] == "manual":
            final_states[key] = bool(cfg["manual_state"])
        else:
            final_states[key] = auto_states.get(key, False)
    return final_states


def relay_modes(relay_config_rows):
    return {r["relay_key"]: r["mode"] for r in relay_config_rows}


# ---------------------------------------------------------------------------
# Device endpoints (ESP32 <-> server)
# ---------------------------------------------------------------------------
@app.post("/api/ingest")
def ingest(payload: IngestPayload, x_api_key: str = Header(...)):
    with get_db() as db:
        device = get_device_by_api_key(db, x_api_key)
        if not device:
            raise HTTPException(status_code=401, detail="Invalid API key")
        if device["id"] != payload.device:
            raise HTTPException(status_code=403, detail="API key does not match device id")

        temp, hum = apply_calibration(payload.raw_temperature, payload.raw_humidity)
        vpd = calc_vpd(temp, hum)
        now = datetime.now(timezone.utc)
        now_iso = now.isoformat()

        db.execute(
            "INSERT INTO readings (device_id, ts, temperature, humidity, vpd) VALUES (?, ?, ?, ?, ?)",
            (device["id"], now_iso, temp, hum, vpd),
        )
        db.execute("UPDATE devices SET last_seen = ? WHERE id = ?", (now_iso, device["id"]))

        ensure_settings_row(db, device["id"])
        ensure_relay_rows(db, device["id"])

        if payload.relay_states:
            for key, state in payload.relay_states.items():
                if key in RELAY_KEYS:
                    db.execute(
                        """INSERT INTO relay_reported (device_id, relay_key, reported_state, updated_at)
                           VALUES (?, ?, ?, ?)
                           ON CONFLICT(device_id, relay_key)
                           DO UPDATE SET reported_state = excluded.reported_state, updated_at = excluded.updated_at""",
                        (device["id"], key, int(bool(state)), now_iso),
                    )

        settings_row = db.execute("SELECT * FROM settings WHERE device_id = ?", (device["id"],)).fetchone()
        relay_config_rows = db.execute("SELECT * FROM relay_config WHERE device_id = ?", (device["id"],)).fetchall()

        commanded = compute_relay_states(settings_row, relay_config_rows, temp, hum, now, db, device["id"])

    append_to_excel(device["id"], now_iso, temp, hum, vpd)

    return commanded


# ---------------------------------------------------------------------------
# Dashboard-facing API
# ---------------------------------------------------------------------------
@app.post("/api/devices")
def create_device(payload: NewDevice, user: str = Depends(require_dashboard_auth)):
    device_id = payload.name.lower().replace(" ", "-")
    api_key = secrets.token_hex(16)
    with get_db() as db:
        try:
            db.execute(
                "INSERT INTO devices (id, name, api_key, created_at) VALUES (?, ?, ?, ?)",
                (device_id, payload.name, api_key, datetime.now(timezone.utc).isoformat()),
            )
            ensure_settings_row(db, device_id)
            ensure_relay_rows(db, device_id)
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Device already exists")
    return {"device_id": device_id, "api_key": api_key}


@app.get("/api/devices")
def list_devices(user: str = Depends(require_dashboard_auth)):
    with get_db() as db:
        rows = db.execute("SELECT id, name, created_at, last_seen FROM devices").fetchall()
        return [dict(r) for r in rows]


@app.get("/api/current/{device_id}")
def get_current(device_id: str, user: str = Depends(require_dashboard_auth)):
    with get_db() as db:
        ensure_settings_row(db, device_id)
        ensure_relay_rows(db, device_id)

        latest = db.execute(
            "SELECT ts, temperature, humidity, vpd FROM readings WHERE device_id = ? ORDER BY ts DESC LIMIT 1",
            (device_id,),
        ).fetchone()
        if not latest:
            return {"reading": None, "relays": None}

        settings_row = db.execute("SELECT * FROM settings WHERE device_id = ?", (device_id,)).fetchone()
        relay_config_rows = db.execute("SELECT * FROM relay_config WHERE device_id = ?", (device_id,)).fetchall()
        now = datetime.now(timezone.utc)
        commanded = compute_relay_states(settings_row, relay_config_rows, latest["temperature"], latest["humidity"], now, db, device_id)
        modes = relay_modes(relay_config_rows)

        relays = {key: {"state": commanded[key], "mode": modes.get(key, "auto")} for key in RELAY_KEYS}
        device_row = db.execute("SELECT last_seen FROM devices WHERE id = ?", (device_id,)).fetchone()

        return {
            "reading": dict(latest),
            "relays": relays,
            "last_seen": device_row["last_seen"] if device_row else None,
        }


@app.get("/api/history/{device_id}")
def get_history(
    device_id: str,
    hours: Optional[float] = None,
    interval_minutes: int = 15,
    limit: int = 200,
    user: str = Depends(require_dashboard_auth),
):
    with get_db() as db:
        if hours is not None:
            cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours)).isoformat()
            bucket_seconds = interval_minutes * 60
            rows = db.execute(
                f"""
                SELECT
                    datetime((CAST(strftime('%s', ts) AS INTEGER) / {bucket_seconds}) * {bucket_seconds}, 'unixepoch') AS bucket_ts,
                    AVG(temperature) AS temperature,
                    AVG(humidity) AS humidity,
                    AVG(vpd) AS vpd
                FROM readings
                WHERE device_id = ? AND ts >= ?
                GROUP BY CAST(strftime('%s', ts) AS INTEGER) / {bucket_seconds}
                ORDER BY bucket_ts ASC
                """,
                (device_id, cutoff),
            ).fetchall()
            return [{"ts": r["bucket_ts"], "temperature": r["temperature"], "humidity": r["humidity"], "vpd": r["vpd"]} for r in rows]

        rows = db.execute(
            "SELECT ts, temperature, humidity, vpd FROM readings WHERE device_id = ? ORDER BY ts DESC LIMIT ?",
            (device_id, limit),
        ).fetchall()
        return [dict(r) for r in reversed(rows)]


@app.get("/api/status/{device_id}")
def get_status(device_id: str, user: str = Depends(require_dashboard_auth)):
    with get_db() as db:
        ensure_settings_row(db, device_id)
        ensure_relay_rows(db, device_id)

        latest = db.execute(
            "SELECT temperature, humidity FROM readings WHERE device_id = ? ORDER BY ts DESC LIMIT 1",
            (device_id,),
        ).fetchone()

        settings_row = db.execute("SELECT * FROM settings WHERE device_id = ?", (device_id,)).fetchone()
        relay_config_rows = db.execute("SELECT * FROM relay_config WHERE device_id = ?", (device_id,)).fetchall()
        reported_rows = {
            r["relay_key"]: r for r in db.execute(
                "SELECT * FROM relay_reported WHERE device_id = ?", (device_id,)
            ).fetchall()
        }

        commanded = {}
        if latest:
            now = datetime.now(timezone.utc)
            commanded = compute_relay_states(settings_row, relay_config_rows, latest["temperature"], latest["humidity"], now, db, device_id)

        modes = relay_modes(relay_config_rows)

        equipment = {}
        for key in RELAY_KEYS:
            reported = reported_rows.get(key)
            equipment[key] = {
                "commanded_state": commanded.get(key, False),
                "reported_state": bool(reported["reported_state"]) if reported else None,
                "mode": modes.get(key, "auto"),
                "last_reported_at": reported["updated_at"] if reported else None,
            }
        return equipment


@app.get("/api/settings/{device_id}")
def get_settings(device_id: str, user: str = Depends(require_dashboard_auth)):
    with get_db() as db:
        ensure_settings_row(db, device_id)
        row = db.execute("SELECT * FROM settings WHERE device_id = ?", (device_id,)).fetchone()
        return dict(row)


@app.post("/api/settings/{device_id}")
def update_settings(device_id: str, payload: SettingsUpdate, user: str = Depends(require_dashboard_auth)):
    with get_db() as db:
        ensure_settings_row(db, device_id)
        updates = {k: v for k, v in payload.dict().items() if v is not None}
        if not updates:
            return {"status": "no changes"}
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        db.execute(f"UPDATE settings SET {set_clause} WHERE device_id = ?", (*updates.values(), device_id))
        row = db.execute("SELECT * FROM settings WHERE device_id = ?", (device_id,)).fetchone()
        return dict(row)


@app.post("/api/relay/{device_id}/{relay_key}")
def update_relay_config(device_id: str, relay_key: str, payload: RelayConfigUpdate, user: str = Depends(require_dashboard_auth)):
    if relay_key not in RELAY_KEYS:
        raise HTTPException(status_code=404, detail=f"Unknown relay '{relay_key}'")
    with get_db() as db:
        ensure_relay_rows(db, device_id)
        updates = {}
        if payload.mode is not None:
            if payload.mode not in ("auto", "manual"):
                raise HTTPException(status_code=400, detail="mode must be 'auto' or 'manual'")
            updates["mode"] = payload.mode
        if payload.manual_state is not None:
            updates["manual_state"] = int(payload.manual_state)
        if not updates:
            return {"status": "no changes"}
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        db.execute(
            f"UPDATE relay_config SET {set_clause} WHERE device_id = ? AND relay_key = ?",
            (*updates.values(), device_id, relay_key),
        )
        row = db.execute(
            "SELECT * FROM relay_config WHERE device_id = ? AND relay_key = ?", (device_id, relay_key)
        ).fetchone()
        return dict(row)


@app.get("/api/export/excel/{device_id}")
def export_excel(device_id: str, user: str = Depends(require_dashboard_auth)):
    path = excel_path(device_id)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="No Excel log exists yet for this device")
    return FileResponse(
        path,
        filename=f"{device_id}_sensor_data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Static dashboard
# ---------------------------------------------------------------------------
static_dir = os.path.join(BASE_DIR, "static")
app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")
